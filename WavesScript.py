from openpyxl import load_workbook
from openpyxl import Workbook



globalWave = load_workbook(filename = "PasswordlessJourney.xlsx")
argentinaWave = Workbook()
chileWave = Workbook()
colombiaWave = Workbook()
costaRicaWave = Workbook()

wsArgentina = argentinaWave.active
wsChile = chileWave.active
wsColombia = colombiaWave.active
wsCostaRica = costaRicaWave.active


def createTable (wsName, wsGlobal):
    wsName["A1"].value = wsGlobal["B1"].value
    wsName["B1"].value = wsGlobal["Q1"].value
    wsName["C1"].value = wsGlobal["R1"].value
    wsName["D1"].value = wsGlobal["T1"].value
    wsName["E1"].value = wsGlobal["W1"].value
    wsName["F1"].value = wsGlobal["AA1"].value

def copyEntireRow (rowNumberOrigin, rowNumber, wsName, wsGlobal):
    wsName["A"+str(rowNumber)].value = wsGlobal["A"+str(rowNumberOrigin)].value
    wsName["B"+str(rowNumber)].value = wsGlobal["Q"+str(rowNumberOrigin)].value
    wsName["C"+str(rowNumber)].value = wsGlobal["R"+str(rowNumberOrigin)].value
    wsName["D"+str(rowNumber)].value = wsGlobal["T"+str(rowNumberOrigin)].value
    wsName["E"+str(rowNumber)].value = wsGlobal["W"+str(rowNumberOrigin)].value
    wsName["F"+str(rowNumber)].value = wsGlobal["AA"+str(rowNumberOrigin)].value


    


#ITERATE THE ORIGINAL EXCEL
sheetNumber = 0
for sheet in globalWave:
    
    wsGlobalName= globalWave.sheetnames[sheetNumber]

    wsGlobal = globalWave[wsGlobalName]

    #RECORRE TODAS LAS FILAS DE UN SHEET. (INVESTIGAR SI SE PUEDE FILTRAR) 
    sourceRow = 2
    desArgRow = 2
    desChiRow = 2
    desColRow = 2
    desCRICARow = 2

    for row in wsGlobal.iter_rows():
        
        if wsGlobal["A1"].value:

            if wsGlobal["C"+str(sourceRow)].value == "Argentina":
                wsArgentinaName = wsArgentina.title
                if (wsGlobalName != wsArgentinaName):
                    argentinaWave.create_sheet(title = wsGlobalName)
                    wsArgentinaName = argentinaWave[wsGlobalName].title
       
                    wsArgentina = argentinaWave[wsArgentinaName]
                    createTable(wsArgentina, wsGlobal)
                copyEntireRow(sourceRow, desArgRow, wsArgentina, wsGlobal)
                desArgRow = desArgRow + 1



            
            elif wsGlobal["C"+str(sourceRow)].value == "Chile":
                wsChileName = wsChile.title

                if (wsGlobalName != wsChileName):
                    chileWave.create_sheet(title = wsGlobalName)
                    wsChileName = chileWave[wsGlobalName].title
                    wsChile = chileWave[wsChileName]
                    createTable(wsChile, wsGlobal)
                copyEntireRow(sourceRow, desChiRow, wsChile, wsGlobal)
                desChiRow = desChiRow + 1
            
            elif wsGlobal["C"+str(sourceRow)].value == "Colombia":
                wsColombiaName = wsColombia.title

                if (wsGlobalName != wsColombiaName):
                    colombiaWave.create_sheet(title = wsGlobalName)
                    wsColombiaName = colombiaWave[wsGlobalName].title
                    wsColombia = colombiaWave[wsColombiaName]
                    createTable(wsColombia, wsGlobal)
                copyEntireRow(sourceRow, desColRow, wsColombia, wsGlobal)
                desColRow = desColRow + 1

            elif wsGlobal["C"+str(sourceRow)].value == "Costa Rica":
                wsCostaRicaName = wsCostaRica.title
                
                if  (wsGlobalName != wsCostaRicaName):
                    costaRicaWave.create_sheet(title = wsGlobalName)
                    wsCostaRicaName = costaRicaWave[wsGlobalName].title
                    wsCostaRica = costaRicaWave[wsCostaRicaName]
                    createTable(wsCostaRica, wsGlobal)
                
                copyEntireRow(sourceRow, desCRICARow, wsCostaRica, wsGlobal)
                desCRICARow = desCRICARow + 1
        sourceRow = sourceRow + 1
    sheetNumber = sheetNumber + 1
    

del argentinaWave["Sheet"]
del chileWave["Sheet"]
del colombiaWave["Sheet"]
del costaRicaWave["Sheet"]

argentinaWave.save("ArgentinaWave.xlsx")
chileWave.save("ChileWave.xlsx")
colombiaWave.save("ColombiaWave.xlsx")
costaRicaWave.save("CostaRicaWave.xlsx")
        







        

                
            